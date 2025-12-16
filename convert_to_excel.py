"""
Convert test results to clean Excel format
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the CSV
df = pd.read_csv('sfa_test_results.csv')

# Clean up response_preview (remove newlines for cleaner Excel)
df['response_preview'] = df['response_preview'].str.replace('\n', ' ').str.replace('  ', ' ')

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "SFA Test Results"

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
headers = ["ID", "Category", "Query", "Expected Source", "Status", "Response"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Write data
for row_idx, row in enumerate(df.itertuples(), 2):
    ws.cell(row=row_idx, column=1, value=row.id).border = thin_border
    ws.cell(row=row_idx, column=2, value=row.category).border = thin_border
    ws.cell(row=row_idx, column=3, value=row.query).border = thin_border
    ws.cell(row=row_idx, column=4, value=row.expected_source).border = thin_border
    
    status_cell = ws.cell(row=row_idx, column=5, value=row.status)
    status_cell.border = thin_border
    if row.status == "PASS":
        status_cell.fill = pass_fill
    
    ws.cell(row=row_idx, column=6, value=row.response_preview).border = thin_border

# Set column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 80

# Wrap text in response column
for row in range(2, len(df) + 2):
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')

# Add summary at bottom
summary_row = len(df) + 3
ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
ws.cell(row=summary_row + 1, column=1, value="Total Queries:")
ws.cell(row=summary_row + 1, column=2, value=len(df))
ws.cell(row=summary_row + 2, column=1, value="Passed:")
ws.cell(row=summary_row + 2, column=2, value=len(df[df['status'] == 'PASS']))
ws.cell(row=summary_row + 3, column=1, value="Accuracy:")
ws.cell(row=summary_row + 3, column=2, value=f"{len(df[df['status'] == 'PASS']) / len(df) * 100:.1f}%")

# Save
output_file = 'sfa_test_results.xlsx'
wb.save(output_file)
print(f"✅ Excel file saved: {output_file}")
