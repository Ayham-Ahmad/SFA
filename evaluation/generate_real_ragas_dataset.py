import os
import time
import pandas as pd
import asyncio
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Ragas Imports
# Ragas Imports
try:
    from ragas import evaluate
    from ragas.metrics import (
        faithfulness,
        answer_relevancy,
        context_precision,
        context_recall
    )
    # Dictionary to hold metrics to use
    metrics_to_use = [faithfulness, answer_relevancy, context_precision, context_recall]
    
    # Try to get context_relevance (deprecated in some versions)
    try:
        from ragas.metrics import context_relevance
        metrics_to_use.append(context_relevance)
    except ImportError:
        try:
            from ragas.metrics import ContextRelevance
            metrics_to_use.append(ContextRelevance())
        except ImportError:
            print("Warning: context_relevance metric not found, skipping.")
            
    from ragas.llms import LangchainLLMWrapper
    from datasets import Dataset
except ImportError as e:
    print(f"CRITICAL ERROR: 'ragas' import failed: {e}")
    exit(1)

# LangChain / LLM Imports
try:
    from langchain_openai import ChatOpenAI
    # Try importing HuggingFaceEmbeddings from new package, fall back to community
    try:
        from langchain_huggingface import HuggingFaceEmbeddings
    except ImportError:
        from langchain_community.embeddings import HuggingFaceEmbeddings
except ImportError:
    print("CRITICAL ERROR: LangChain dependencies missing. Run: pip install langchain-openai langchain-community sentence-transformers")
    exit(1)

# Backend Imports (for live logic)
from backend.agents.planner import plan_task
from backend.agents.worker import execute_step
from backend.agents.auditor import audit_and_synthesize
from backend.security.safety import sanitize_content
from dotenv import load_dotenv

# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------
load_dotenv()
GROQ_API_KEY = os.getenv("GROQ_API_KEY")

if not GROQ_API_KEY:
    print("ERROR: GROQ_API_KEY not found in .env")
    exit(1)

# Initialize Judge LLM (Groq via OpenAI Protocol)
# This saves money/tokens by using Groq instead of GPT-4, assuming user wants to use their keys.
judge_llm = ChatOpenAI(
    base_url="https://api.groq.com/openai/v1",
    api_key=GROQ_API_KEY,
    model="llama-3.3-70b-versatile", # Strong model for judging
    temperature=0
)

# Initialize Embeddings
# Must match what is likely used or be a good standard.
print("Initializing Embeddings model (this may take a moment)...")
embeddings = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")

# -----------------------------------------------------------------------------
# Test Data (Ground Truths)
# -----------------------------------------------------------------------------
test_queries = [
    {
        "question": "What was the total revenue for Apple Inc in 2022?",
        "ground_truth": "Apple Inc. reported total revenue of approximately $394.3 billion in 2022."
    },
    {
        "question": "Compare the Net Income of Microsoft and Apple for the year 2022.",
        "ground_truth": "In 2022, Apple's Net Income was ~$99.8B while Microsoft's was ~$72.7B."
    },
    {
        "question": "What are the total assets held by Tesla, Inc. as of the latest report?",
        "ground_truth": "Tesla's total assets were approximately $82.3 billion in 2022."
    },
    {
        "question": "Plot the revenue trend for Amazon from 2020 to 2023.",
        "ground_truth": "Amazon revenue trend: 2020: $386B, 2021: $469B, 2022: $514B. Trend is consistently increasing."
    },
    {
        "question": "What is the debt-to-equity ratio for Netflix Inc in 2022?",
        "ground_truth": "Netflix Total Liabilities: ~$30B. Equity: ~$20B. D/E Ratio approx 1.5."
    },
    {
        "question": "List the top 3 separate expense categories for Nvidia Corp.",
        "ground_truth": "Major expenses for Nvidia typically include Research & Development and Cost of Revenue."
    },
    {
        "question": "Did Apple Inc report a profit or loss in 2023?",
        "ground_truth": "Apple reported a profit (Net Income) in 2023."
    },
    {
        "question": "What is the Gross Profit margin for Microsoft in 2022?",
        "ground_truth": "Microsoft Gross Margin is approx 68% (Revenue $198B, Gross Profit $135B)."
    },
    {
        "question": "How much cash did Tesla have on hand at the end of 2022?",
        "ground_truth": "Tesla reported Cash and Cash Equivalents of approximately $16.2 billion."
    },
    {
        "question": "What are the primary risk factors for Apple Inc?",
        "ground_truth": "Supply chain risks, reliance on iPhone sales, and regulatory scrutiny in various markets."
    },
    {
        "question": "Compare the revenue growth of Nvidia vs Amazon in 2022.",
        "ground_truth": "Amazon grew ~9%. Nvidia grew significantly faster but faced cyclical headwinds in gaming."
    },
    {
        "question": "Show the balance sheet summary for Meta Platforms.",
        "ground_truth": "Meta (2022): Assets ~$185B, Liabilities ~$60B, Equity ~$125B."
    },
    {
        "question": "What was the highest quarterly revenue for Apple in 2022?",
        "ground_truth": "Q1 2022 (Holiday Quarter) usually shows the highest revenue, ~123.9B."
    },
    {
        "question": "Does Microsoft pay dividends? If so, how much?",
        "ground_truth": "Yes, Microsoft pays dividends. Example: $0.62 per share quarterly (approx)."
    },
    {
        "question": "What is the total Stockholders Equity for Amazon?",
        "ground_truth": "Amazon Stockholders Equity was approximately $146 billion in 2022."
    },
    {
        "question": "Rank Apple, Microsoft, and Tesla by total assets in 2022.",
        "ground_truth": "1. Microsoft (~$364B), 2. Apple (~$352B), 3. Tesla (~$82B)."
    },
    {
        "question": "What is the EPS (Earnings Per Share) for Nvidia in 2022?",
        "ground_truth": "Nvidia EPS was approximately $3.85 (split adjusted)."
    },
    {
        "question": "Analyze the cash flow from operations for Netflix.",
        "ground_truth": "Netflix Operating Cash Flow has turned positive, approx $2B in 2022."
    },
    {
        "question": "What are the total liabilities for 10X Genomics?",
        "ground_truth": "Total liabilities for 10X Genomics were approx $200M-$400M range depending on quarter."
    },
    {
        "question": "Is Tesla's revenue greater than its expenses?",
        "ground_truth": "Yes, Tesla is profitable, meaning Revenue > Expenses (Net Income is positive)."
    }
]

# -----------------------------------------------------------------------------
# Pipeline Execution (Replica of run_ramas_pipeline with Context Capture)
# -----------------------------------------------------------------------------
def run_pipeline_with_capture(question: str):
    """
    Runs the SFA agents but captures retrieved contexts for Ragas.
    """
    print(f"Propagating: {question}")
    
    # 1. PLANNER
    try:
        plan = plan_task(question)
    except Exception as e:
        return "Error in planning", [str(e)]

    contexts = []
    
    # 2. WORKER
    full_context_str = ""
    steps = plan.split("\n")
    for step in steps:
        if step.strip() and (step[0].isdigit() or step.strip().startswith("-")):
            clean_step = step.replace("**", "")
            if "." in clean_step:
                clean_step = clean_step.split(".", 1)[1].strip()
            
            try:
                result = execute_step(clean_step)
                # Store strict context (the result of the tool)
                contexts.append(result)
                full_context_str += f"\nStep: {step}\nResult: {result}\n"
            except Exception as e:
                contexts.append(f"Error executing step '{clean_step}': {e}")

    # 3. AUDITOR (Custom Evaluation Mode)
    try:
        from groq import Groq
        client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
        
        evaluation_audit_prompt = f"""
        You are an evaluator in a RAG system.
        
        GOAL: Answer the User Question strictly based on the Retrieved Context.
        
        User Question: {question}
        
        Retrieved Context:
        {full_context_str}
        
        CRITICAL RULES:
        1. Answer in PLAIN ENGLISH TEXT consistent with the Ground Truth style.
        2. Do NOT format as a dashboard, no graphs, no SQL blocks.
        3. If numbers are present, state them clearly (e.g., "$500 million").
        4. If the context has no relevant data, say "Data not available in context."
        
        Refrain from polite conversation. Just give the factual answer.
        """
        
        resp = client.chat.completions.create(
            messages=[{"role": "user", "content": evaluation_audit_prompt}],
            model="llama-3.3-70b-versatile",
            temperature=0
        )
        final_answer = resp.choices[0].message.content.strip()

    except Exception as e:
        final_answer = f"Error generating answer: {e}"
        
    return final_answer, contexts

# -----------------------------------------------------------------------------
# Main Generation Loop
# -----------------------------------------------------------------------------
def main():
    print(f"\n--- Starting RAGAS Dataset Generation ---")
    print(f"Loaded {len(test_queries)} test cases.")
    
    results_data = {
        "question": [],
        "answer": [],
        "contexts": [],
        "ground_truth": []
    }
    
    # 1. Generate Live Data
    # Initialize Excel Dataframe
    output_file = "ragas_evaluation_SFA_v2.xlsx"
    processed_count = 0
    df_intermediate = pd.DataFrame(columns=[
        "Query", "Ground Truth", "Retrieved Context", "Model Answer",
        "Faithfulness", "Ans Rel", "Ctx Prec", "Ctx Recall", "Ctx Rel"
    ])

    # Check for existing progress to resume
    if os.path.exists(output_file):
        try:
            existing_df = pd.read_excel(output_file)
            # Basic validation to ensure it has the right columns
            if "Query" in existing_df.columns and len(existing_df) > 0:
                print(f"Resuming from existing file with {len(existing_df)} entries.")
                df_intermediate = existing_df
                processed_count = len(existing_df)
                
                # Re-populate results lists so Ragas has full dataset at the end
                results_data['question'] = df_intermediate['Query'].tolist()
                results_data['ground_truth'] = df_intermediate['Ground Truth'].tolist()
                results_data['answer'] = df_intermediate['Model Answer'].tolist()
                # Context is stored as string in Excel, try to restore to list or keep as list-in-string
                # For simplicity in resume, we might just put the string back in a list
                results_data['contexts'] = [[x] for x in df_intermediate['Retrieved Context'].tolist()]
            else:
                 print("Existing file found but empty or invalid format. Starting fresh.")
        except Exception as e:
            print(f"Error reading existing file: {e}. Starting fresh.")
    
    for i, item in enumerate(test_queries):
        if i < processed_count:
            print(f"[{i+1}/{len(test_queries)}] Already processed: {item['question']}")
            continue

        q = item['question']
        gt = item['ground_truth']
        
        
        print(f"\n[{i+1}/{len(test_queries)}] Processing: {q}")
        
        try:
            # Add timeout protection manually if possible, or just catch all
            answer, retrieved_contexts = run_pipeline_with_capture(q)
        except Exception as e:
            print(f"Error processing query: {e}")
            answer = "ERROR"
            retrieved_contexts = [str(e)]
        
        # Format Context
        ctx_str = "\n".join(retrieved_contexts)[:3000] if retrieved_contexts else "No context."
        
        # Add to temporary DF
        new_row = {
            "Query": q,
            "Ground Truth": gt,
            "Retrieved Context": ctx_str,
            "Model Answer": answer,
            "Faithfulness": 0.0,
            "Ans Rel": 0.0,
            "Ctx Prec": 0.0,
            "Ctx Recall": 0.0,
            "Ctx Rel": 0.0
        }
        
        # Append to master lists for Ragas later
        results_data['question'].append(q)
        results_data['answer'].append(answer)
        results_data['contexts'].append(retrieved_contexts if retrieved_contexts else ["No context."])
        results_data['ground_truth'].append(gt)
        
        # Incremental Save to Excel (so we see progress)
        df_intermediate = pd.concat([df_intermediate, pd.DataFrame([new_row])], ignore_index=True)
        df_intermediate.to_excel(output_file, index=False)
        print(f"Saved progress to {output_file}")
        
        # Brief pause
        time.sleep(1)

    # 2. Convert to Dataset
    dataset = Dataset.from_dict(results_data)
    
    print("\n--- Starting Ragas Evaluation (This consumes API credits) ---")
    
    # 3. Evaluate
    # We use a custom run configuration to avoid hitting rate limits if possible
    try:
        evaluation_results = evaluate(
            dataset=dataset,
            metrics=metrics_to_use,
            llm=judge_llm,
            embeddings=embeddings,
            raise_exceptions=False # Don't crash entire batch on one error
        )
        
        df_results = evaluation_results.to_pandas()
        
    except Exception as e:
        print(f"\nError during Ragas Evaluation: {e}")
        # Fallback: Save what we have without metrics
        df_results = pd.DataFrame(results_data)
        for metric in ['faithfulness', 'answer_relevancy', 'context_precision', 'context_recall', 'context_relevance']:
            df_results.loc[:, metric] = 0.0

    # 4. Save to Excel with Formatting
    print("\n--- Saving Results to Excel ---")
    output_file = "ragas_evaluation_SFA.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "RAGAS Evaluation"
    
    # Reorder columns for readability
    desired_order = [
        'question', 'ground_truth', 'contexts', 'answer',
        'faithfulness', 'answer_relevancy', 'context_precision', 
        'context_recall', 'context_relevance'
    ]
    
    # Ensure columns exist (if Ragas failed, some might be missing)
    for col in desired_order:
        if col not in df_results.columns:
            df_results[col] = 0.0
            
    # Select and rename for report
    final_df = df_results[desired_order].copy()
    final_df.columns = [
        "Query", "Ground Truth", "Retrieved Context", "Model Answer",
        "Faithfulness", "Ans Rel", "Ctx Prec", "Ctx Recall", "Ctx Rel"
    ]
    
    # Convert 'Retrieved Context' list to string for Excel
    final_df["Retrieved Context"] = final_df["Retrieved Context"].apply(lambda x: "\n".join(x)[:3000] if isinstance(x, list) else str(x)[:3000])

    # Write Header
    headers = list(final_df.columns)
    ws.append(headers)
    
    # Write Rows
    for r in dataframe_to_rows(final_df, index=False, header=False):
        ws.append(r)
        
    # Styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Column Widths
    dims = {'A': 40, 'B': 40, 'C': 50, 'D': 40} # Text columns
    for col, width in dims.items():
        ws.column_dimensions[col].width = width
        
    # Metric Cols
    for col in ['E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12
        
    # Wrap Text
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

    wb.save(output_file)
    print(f"successfully saved to {output_file}")
    print(f"Total Cost Estimate: ~{len(test_queries) * 5} API calls made.")

if __name__ == "__main__":
    main()
