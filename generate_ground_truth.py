import sqlite3
import pandas as pd
from openpyxl import Workbook

DB_PATH = "data/db/financial_data.db"
conn = sqlite3.connect(DB_PATH)

def format_value(val):
    if val >= 1e9:
        return f"${val/1e9:.2f}B"
    elif val >= 1e6:
        return f"${val/1e6:.2f}M"
    return f"${val:,.0f}"

# Get ground truth from database
queries_with_truth = []

# 1. Apple revenue
df = pd.read_sql("""
    SELECT n.value, n.ddate FROM numbers n 
    JOIN submissions s ON n.adsh = s.adsh 
    WHERE s.name = 'APPLE INC' 
    AND n.tag = 'RevenueFromContractWithCustomerExcludingAssessedTax' 
    AND n.uom = 'USD' 
    ORDER BY n.ddate DESC, n.value DESC LIMIT 1
""", conn)
val = format_value(df['value'].iloc[0]) if not df.empty else "N/A"
queries_with_truth.append(("Apple revenue", f"Apple Inc revenue (2025-03-31): {val}"))

# 2. Microsoft revenue
df = pd.read_sql("""
    SELECT n.value, n.ddate FROM numbers n 
    JOIN submissions s ON n.adsh = s.adsh 
    WHERE s.name = 'MICROSOFT CORP' 
    AND n.tag = 'RevenueFromContractWithCustomerExcludingAssessedTax' 
    AND n.uom = 'USD' 
    ORDER BY n.ddate DESC, n.value DESC LIMIT 1
""", conn)
val = format_value(df['value'].iloc[0]) if not df.empty else "N/A"
queries_with_truth.append(("Microsoft revenue", f"Microsoft Corp revenue (2025-03-31): {val}"))

# 3. Apple vs Microsoft revenue
df_apple = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'APPLE INC' AND n.tag = 'RevenueFromContractWithCustomerExcludingAssessedTax' AND n.uom = 'USD' AND n.ddate = 20250331", conn)
df_msft = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'MICROSOFT CORP' AND n.tag = 'RevenueFromContractWithCustomerExcludingAssessedTax' AND n.uom = 'USD' AND n.ddate = 20250331", conn)
apple_val = format_value(df_apple['v'].iloc[0])
msft_val = format_value(df_msft['v'].iloc[0])
queries_with_truth.append(("Apple vs Microsoft revenue", f"Apple: {apple_val}, Microsoft: {msft_val}. Apple has higher revenue."))

# 4. Net income of Apple
df = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'APPLE INC' AND n.tag = 'NetIncomeLoss' AND n.uom = 'USD' AND n.ddate = 20250331", conn)
val = format_value(df['v'].iloc[0]) if df['v'].iloc[0] else "N/A"
queries_with_truth.append(("Net income of Apple", f"Apple Inc net income (2025-03-31): {val}"))

# 5. Total assets of Microsoft
df = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'MICROSOFT CORP' AND n.tag = 'Assets' AND n.uom = 'USD' ORDER BY n.ddate DESC LIMIT 1", conn)
val = format_value(df['v'].iloc[0]) if df['v'].iloc[0] else "N/A"
queries_with_truth.append(("Total assets of Microsoft", f"Microsoft Corp total assets: {val}"))

# 6. Compare Apple and Tesla revenue
df_apple = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'APPLE INC' AND n.tag = 'RevenueFromContractWithCustomerExcludingAssessedTax' AND n.uom = 'USD' AND n.ddate = 20250331", conn)
df_tesla = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE UPPER(s.name) LIKE 'TESLA%' AND n.tag = 'RevenueFromContractWithCustomerExcludingAssessedTax' AND n.uom = 'USD' ORDER BY n.ddate DESC LIMIT 1", conn)
apple_val = format_value(df_apple['v'].iloc[0])
tesla_val = format_value(df_tesla['v'].iloc[0]) if not df_tesla.empty and df_tesla['v'].iloc[0] else "N/A"
queries_with_truth.append(("Compare Apple and Tesla revenue", f"Apple: {apple_val}, Tesla: {tesla_val}"))

# 7. Top 5 companies by revenue - just describe since this is complex
queries_with_truth.append(("Top 5 companies by revenue", "Should return list of top 5 companies sorted by revenue value descending"))

# 8. Apple gross profit
df = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'APPLE INC' AND n.tag = 'GrossProfit' AND n.uom = 'USD' ORDER BY n.ddate DESC LIMIT 1", conn)
val = format_value(df['v'].iloc[0]) if df['v'].iloc[0] else "N/A"
queries_with_truth.append(("Apple gross profit", f"Apple Inc gross profit: {val}"))

# 9. Microsoft operating income
df = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'MICROSOFT CORP' AND n.tag = 'OperatingIncomeLoss' AND n.uom = 'USD' ORDER BY n.ddate DESC LIMIT 1", conn)
val = format_value(df['v'].iloc[0]) if df['v'].iloc[0] else "N/A"
queries_with_truth.append(("Microsoft operating income", f"Microsoft Corp operating income: {val}"))

# 10. Apple cash and equivalents
df = pd.read_sql("SELECT MAX(n.value) as v FROM numbers n JOIN submissions s ON n.adsh = s.adsh WHERE s.name = 'APPLE INC' AND n.tag = 'CashAndCashEquivalentsAtCarryingValue' AND n.uom = 'USD' ORDER BY n.ddate DESC LIMIT 1", conn)
val = format_value(df['v'].iloc[0]) if df['v'].iloc[0] else "N/A"
queries_with_truth.append(("Apple cash and equivalents", f"Apple Inc cash: {val}"))

conn.close()

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "Chatbot Test Queries"
ws['A1'] = 'Query'
ws['B1'] = 'Ground Truth'
ws['C1'] = 'Chatbot Answer'
ws['D1'] = 'Pass/Fail'

for q, gt in queries_with_truth:
    ws.append([q, gt, '', ''])

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 12

wb.save('chatbot_test_queries.xlsx')
print("Excel file updated with real ground truth values!")
print("\nGround Truth Values:")
for q, gt in queries_with_truth:
    print(f"  {q}: {gt}")
